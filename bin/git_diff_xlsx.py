# Converts a Microsoft Excel 2007+ file into plain text
# for comparison using git diff
#
# Instructions for setup:
# 1. Place this file in a folder
# 2. Add the following line to the global .gitconfig:
#     [diff "zip"]
#   	    binary = True
#	    textconv = python c:/path/to/git_diff_xlsx.py
# 3. Add the following line to the repository's .gitattributes
#    *.xlsx diff=zip
# 4. Now, typing [git diff] at the prompt will produce text versions
# of Excel .xlsx files
#
# Copyright William Usher 2013
# Contact: w.usher@ucl.ac.uk
#

import sys
import openpyxl as xl
from openpyxl import load_workbook

def parse(infile,outfile):
    """
    Converts an Excel file into text
    Returns a formatted text file for comparison using git diff.
    """

    book = xl.load_workbook(infile)

    num_sheets = book.get_sheet_names()

    print num_sheets

#   print "File last edited by " + book.user_name + "\n"
    if ( book.properties.lastModifiedBy ):
        outfile.write("File last edited by " + book.properties.lastModifiedBy + "\n")

    # loop over worksheets

    for index in book.get_sheet_names():
        # find non empty cells
        sheet = book.get_sheet_by_name(index)
        outfile.write("=================================\n")
        outfile.write("Sheet: " + index + "[ " + str(sheet.get_highest_row()) + " , " + str(sheet.get_highest_column()) + " ]\n")
        outfile.write("=================================\n")
        for row in range(1,sheet.get_highest_row()+1):
            for col in range(1,sheet.get_highest_column()+1):
                content = sheet.cell(column=col, row=row).value
                if content <> "":
                    outfile.write("    " + xl.utils.get_column_letter(col) + str(row) + ": " + unicode(content) + "\n")
                    #outfile.write("    " + unicode(xl.utils.get_column_letter(col)) + row + ": " + unicode(content) + "\n")
        print "\n"

# output cell address and contents of cell
def main():
    args = sys.argv[1:]
    if len(args) != 1:
        print 'usage: python git_diff_xlsx.py infile.xlsx'
        sys.exit(-1)
    outfile = sys.stdout
    parse(args[0],outfile)

if __name__ == '__main__':
    main()
